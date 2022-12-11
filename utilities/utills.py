import logging

import inspect
# from numpy import append

from openpyxl import load_workbook

class Utils:
    def logger(log_level = logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(log_level)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s')

        fh = logging.FileHandler(r"D:\IDE Projects\VS Code Projects\QA_Selenium\Assignment_5\logger.log",mode="a")

        fh.setFormatter(formatter)
        logger.addHandler(fh)

        return logger

    def read_data_from_excel(file_name,sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2,row_ct+1):
            row= []
            for j in range(1,col_ct-1):
            
                row.append(sh.cell(row=i,column = j).value)
            row.append(i)
            datalist.append(row)
        print(datalist)
        return datalist

    def write_data_to_excel(file_name,sheet,result,row):
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]        
        sh.cell(row=row,column=4).value=result
        wb.save(file_name)
